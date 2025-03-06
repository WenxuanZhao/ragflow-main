<<<<<<< HEAD
#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
=======
>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

<<<<<<< HEAD
from openpyxl import load_workbook
=======
import logging
from openpyxl import load_workbook, Workbook
>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
import sys
from io import BytesIO

from rag.nlp import find_codec

<<<<<<< HEAD

class RAGFlowExcelParser:
    def html(self, fnm, chunk_rows=256):
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))

=======
import pandas as pd


class RAGFlowExcelParser:
    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        try:
            return load_workbook(file_like_object)
        except Exception as e:
            logging.info(f"****wxy: openpyxl load error: {e}, try pandas instead")
            try:
                df = pd.read_excel(file_like_object)
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                for col_num, column_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_num, value=column_name)
                for row_num, row in enumerate(df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                return wb
            except Exception as e_pandas:
                raise Exception(f"****wxy: pandas read error: {e_pandas}, original openpyxl error: {e}")

    def html(self, fnm, chunk_rows=256):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
        tb_chunks = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(
<<<<<<< HEAD
                    rows[1 + chunk_i * chunk_rows : 1 + (chunk_i + 1) * chunk_rows]
=======
                  rows[1 + chunk_i * chunk_rows: 1 + (chunk_i + 1) * chunk_rows]
>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
                ):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
<<<<<<< HEAD
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))
=======
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)

>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                line = "; ".join(fields)
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
<<<<<<< HEAD
            wb = load_workbook(BytesIO(binary))
=======
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
<<<<<<< HEAD
=======

>>>>>>> 4f9504305a238b4fd3346c988bb1e7872b79d192
